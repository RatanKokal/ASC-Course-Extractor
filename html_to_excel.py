import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook, Workbook
import os
import numpy as np
import re

def create_empty_excel(filename):
    """
    Creates an empty Excel file if it doesn't exist.

    Parameters:
    - filename (str): Name of the Excel file to create.
    """
    if not os.path.exists(filename):
        wb = Workbook()
        sheet = wb.active
        sheet.title = 'Course List'
        wb.save(filename)
        wb.close()

def add_data(next_row, df, ws):
    """
    Adds data from DataFrame to Excel worksheet starting from 'next_row'.

    Parameters:
    - next_row (int): Next row index to start adding data.
    - df (DataFrame): DataFrame containing data to add.
    - ws (Worksheet): Excel worksheet object.
    """
    if next_row == 1:
        # Add column names if the sheet is empty
        for col_index, value in enumerate(df.columns, start=1):
            ws.cell(row=next_row, column=col_index, value=value)
        next_row += 1
    
    # Add data rows
    for index, row in df.iterrows():
        for col_index, value in enumerate(row, start=1):
            if isinstance(value, int):
                ws.cell(row=next_row + index, column=col_index, value=int(value))
            else:
                ws.cell(row=next_row + index, column=col_index, value=value)

    return next_row

def get_maximum_rows(sheet_object):
    """
    Returns the maximum row number containing data in the worksheet.

    Parameters:
    - sheet_object (Worksheet): Excel worksheet object.

    Returns:
    - int: Maximum row number containing data.
    """
    rows = 0
    for max_row, row in enumerate(sheet_object, 1):
        if not all(col.value is None for col in row):
            rows += 1
    return rows

# Main script
if __name__ == "__main__":
    # Ensure the Excel file exists
    create_empty_excel('course.xlsx')

    # Process each HTML file in the current directory
    for filename in os.listdir('./'):
        if filename.endswith('html'):
            with open(filename) as f:
                soup = BeautifulSoup(f, 'html.parser')

                # Extract div elements with IDs starting with 'place' or '1place'
                div_elements = soup.find_all('div')
                restriction_table = {}
                section_table = {}

                for div in div_elements:
                    if 'id' in div.attrs:
                        if div.attrs['id'].startswith('place'):
                            restriction_table[div.attrs['id']] = div.extract()
                        elif div.attrs['id'].startswith('1place'):
                            section_table[div.attrs['id']] = div.extract()

                # Extract table from HTML
                table = soup.find('table')
                rows_to_remove = 3
                for row in table.find_all('tr')[:rows_to_remove]:
                    row.decompose()

                df = pd.read_html(str(table))[0]
                df.columns = df.iloc[0] 
                df = df[1:].reset_index(drop=True)
                df['Sr no.'] = df['Sr no.'].astype(int)
                df['Registration Limit'] = df['Registration Limit'].apply(lambda x: int(x) if not pd.isna(x) else x)
                df['Slot Number'] = df['Slot'].apply(lambda x: x.split()[0] if isinstance(x, str) else None)
                df['Slot Number'] = df['Slot Number'].apply(lambda x: int(x) if x is not None and x.isdigit() else x)
                df['Slot Number'] = df.apply(lambda row: row['Slot'].split('-')[1] if isinstance(row['Slot Number'], str) else row['Slot Number'], axis=1)
                df['Slot Number'] = df['Slot Number'].apply(lambda x: int(re.match(r'^\d+', x).group()) if isinstance(x, str) and re.match(r'^\d+', x) else x)
                df['Department'] = df['Course Code'].apply(lambda x: re.match(r'^[A-Z]+', x).group() if re.match(r'^[A-Z]+', x) else None)

                # Load existing Excel file
                wb = load_workbook('course.xlsx')
                ws = wb['Course List']

                # Find next empty row in Excel sheet
                next_row = get_maximum_rows(ws) + 1

                # Add main data to Excel sheet
                next_row = add_data(next_row, df, ws)

                # Process each restriction table
                for key, value in restriction_table.items():
                    new_df = pd.read_html(str(value))[1]
                    new_df.columns = new_df.iloc[0] 
                    new_df = new_df[1:].reset_index(drop=True)
                    mod_key = key.replace('place', '')
                    title = df.loc[df['Sr no.'] == int(mod_key) + 1, 'Course Code'].values[0] + ' Restrictions'
                    wb.create_sheet(title=title)
                    new_ws = wb[title]
                    add_data(1, new_df, new_ws)
                    # Create hyperlink to the newly added sheet
                    ws.cell(row=next_row + int(mod_key), column=13).hyperlink = f"#'{title}'!A1"

                # Save and close the Excel file
                wb.save('course.xlsx')
                wb.close()
            os.remove(filename)
